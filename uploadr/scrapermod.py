import lxml.html
from lxml.cssselect import CSSSelector
from openpyxl import Workbook
from openpyxl import load_workbook
import textract
import requests
import re
import sys, os

#saving html request to PDF
#http://stackoverflow.com/questions/34503412/download-and-save-pdf-file-with-python-requests-module
#with open('/tmp/metadata.pdf', 'wb') as f:
#    f.write(response.content)

#using textract
#https://media.readthedocs.org/pdf/textract/latest/textract.pdf

#working with excel files in python
#https://openpyxl.readthedocs.io/en/default/

#load queries from input excel worksheet

def scrape_from_pdf(inputfile):
	wb = load_workbook(inputfile, keep_vba=True) 
	ws = wb.active
	queries = []
	for index, row in enumerate(ws.iter_rows()):
		if(index > 0):
			queries.append(row[0].value)

	wb = Workbook()
	ws = wb.active
	ws.append(["Query", "Name", "Title", "Email","Address","Error"])

	for q in queries:
		try:
			search_results = requests.post('https://ecorp.sos.ga.gov/BusinessSearch', data = {'search.SearchType':'BusinessName', 'search.SearchValue':q, 'search.SearchCriteria':"StartsWith"})
			tree = lxml.html.fromstring(search_results.content)
			
			businesses_elements = tree.cssselect('#grid_businessList > tbody > tr > td:nth-child(1) > a')

			try: #check to see if there are any businesses found using the query
				first_business_href = businesses_elements[0].get('href')
				print first_business_href
			except Exception, e:
				print "%s Business -- not found" % q
				ws.append([q,None,None,None,None,"Business Not Found"])
				continue

			businessId = re.findall(r'businessId=(\d+)', first_business_href)[0]
			print businessId

			filings_list = requests.post('https://ecorp.sos.ga.gov/BusinessSearch/BusinessFilings', data = {'businessId':businessId}, headers={'Referer':first_business_href})
			tree = lxml.html.fromstring(filings_list.content)

			most_recent_filings = tree.cssselect('#xhtml_grid > tbody > tr:nth-last-child(1) > td:nth-child(4) > a')
			try:
				most_recent_filing = most_recent_filings[0].get('href')
			except Exception, e:
				print "%s Annual Filing Not Found" % q
				ws.append([q,None,None,None,None,"Annual Filing Not Found"])
				continue

			print most_recent_filing

			filing_no = re.findall(r'filingNo=(\d+)', most_recent_filing)[0]
			print filing_no

			document_id = requests.post('https://ecorp.sos.ga.gov/BusinessSearch/FilingDocuments', data = {'filingNo':filing_no}, headers={'Referer':most_recent_filing}).content

			if (int(document_id) == 0):
				print "%s Annual Filing Not Found" % q
				ws.append([q,None,None,None,None,"Annual Filing Not Found"])
				continue

			get_pdf = requests.post('https://ecorp.sos.ga.gov/BusinessSearch/DownloadFile', data = {'documentId':document_id}, headers={'Referer': 'https://ecorp.sos.ga.gov/BusinessSearch/BusinessFilings'})
			#pdf_filename = "document_id" + '.pdf'
			pdf_filename = 'temp.pdf'
			with open(pdf_filename, 'wb') as f:
				f.write(get_pdf.content)

			text = textract.process(pdf_filename)

			#runs fine locally, but not on remote server. Text extract version?
			# address = re.findall(r'Amount Due AFTER.+\n\n.+?\n((.+?\n){2,3})', text)[0][0]
			# name = re.findall(r"AUTHORIZED SIGNATURE:\n\n(.+?)\n", text)[0]
			# title = re.findall(r'Title:(.+?)\s+', text)[0]
			# email = re.findall(r'([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)', text)[0]

			#runs on server
			address = re.findall(r'Information on record as of.+\n\n.+?\n((.+?\n){2,3})', text)[0][0]
			name = re.findall(r"AUTHORIZED SIGNATURE:\s+(.+?)\n", text)[0]
			title = re.findall(r'Title:(.+?)\s+\n', text)[0]
			email = re.findall(r'([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)', text)[0]

			entry = [q, name,title,email,address]
			print entry
			ws.append(entry)
		except Exception, e:
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			print(exc_type, fname, exc_tb.tb_lineno)
			print q, text
			continue

	wb.save("sample.xlsx")
	return "sample.xlsx"
